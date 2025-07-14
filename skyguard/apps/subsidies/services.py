"""
Subsidies services for the GPS system.
Migrated from legacy django14 system to modern architecture.
"""
import json
import csv
import os
from datetime import datetime, timedelta, date
from typing import Dict, List, Any, Optional, Tuple
from decimal import Decimal
from django.db.models import Q, Sum, Count, Avg
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Driver, DailyLog, CashReceipt, TimeSheetCapture, 
    SubsidyRoute, SubsidyReport, EconomicMapping
)
from skyguard.apps.gps.models import GPSDevice, GPSEvent


# Route configurations migrated from legacy system
RUTA_A6 = [
    'A6 - 01', 'A6 - 02', 'A6 - 03', 'A6 - 04', 'A6 - 05', 'A6 - 06', 'A6 - 07', 'A6 - 08', 'A6 - 09', 'A6 - 10',
    'A6 - 11', 'A6 - 12', 'A6 - 13', 'A6 - 14', 'A6 - 15', 'A6 - 16', 'A6 - 17', 'A6 - 18', 'A6 - 19', 'A6 - 20', 
    'A6 - 21', 'A6 - 22', 'A6 - 26'
]

RUTA_155 = [
    'R155 - 01', 'R155 - 02', 'R155 - 03', 'R155 - 04', 'R155 - 05', 'R155 - 06', 'R155 - 07', 'R155 - 08',
    'R155 - 09', 'R155 - 10', 'R155 - 11', 'R155 - 12', 'R155 - 13', 'R155 - 14', 'R155 - 15', 'R155 - 16',
    'R155 - 17', 'R155 - 18', 'R155 - 19', 'R155 - 20', 'R155 - 21', 'R155 - 22', 'R155 - 23', 'R155 - 24',
    'R155 - 25', 'R155 - 26', 'R155 - 27', 'R155 - 28', 'R155 - 29', 'R155 - 30', 'R155 - 31', 'R155 - 32'
]

RUTA_202 = [
    'R202 - 01', 'R202 - 02', 'R202 - 03', 'R202 - 04', 'R202 - 05', 'R202 - 06', 'R202 - 07', 'R202 - 08',
    'R202 - 09', 'R202 - 10', 'R202 - 11', 'R202 - 12', 'R202 - 13', 'R202 - 14', 'R202 - 15', 'R202 - 16',
    'R202 - 17', 'R202 - 18', 'R202 - 19', 'R202 - 20', 'R202 - 21', 'R202 - 22', 'R202 - 23', 'R202 - 24',
    'R202 - 25', 'R202 - 26', 'R202 - 27', 'R202 - 28', 'R202 - 29', 'R202 - 30', 'R202 - 31', 'R202 - 32'
]

RUTA_31 = [
    'R31 - 01', 'R31 - 02', 'R31 - 03', 'R31 - 04', 'R31 - 05', 'R31 - 06', 'R31 - 07', 'R31 - 08',
    'R31 - 09', 'R31 - 10', 'R31 - 11', 'R31 - 12', 'R31 - 13', 'R31 - 14', 'R31 - 15', 'R31 - 16',
    'R31 - 17', 'R31 - 18', 'R31 - 19', 'R31 - 20', 'R31 - 21', 'R31 - 22', 'R31 - 23', 'R31 - 24'
]

RUTA_400_S1 = [
    'R400 - 01', 'R400 - 02', 'R400 - 03', 'R400 - 04', 'R400 - 05', 'R400 - 06', 'R400 - 07', 'R400 - 08',
    'R400 - 09', 'R400 - 10', 'R400 - 11', 'R400 - 12', 'R400 - 13', 'R400 - 14', 'R400 - 15', 'R400 - 16',
    'R400 - 17', 'R400 - 18', 'R400 - 19', 'R400 - 20', 'R400 - 21', 'R400 - 22', 'R400 - 23', 'R400 - 24',
    'R400 - 25', 'R400 - 26', 'R400 - 27', 'R400 - 28', 'R400 - 29', 'R400 - 30', 'R400 - 31', 'R400 - 32'
]

RUTA_400_S2 = [
    'R400 - 33', 'R400 - 34', 'R400 - 35', 'R400 - 36', 'R400 - 37', 'R400 - 38', 'R400 - 39', 'R400 - 40',
    'R400 - 41', 'R400 - 42', 'R400 - 43', 'R400 - 44', 'R400 - 45', 'R400 - 46', 'R400 - 47', 'R400 - 48',
    'R400 - 49', 'R400 - 50', 'R400 - 51', 'R400 - 52', 'R400 - 53', 'R400 - 54', 'R400 - 55', 'R400 - 56',
    'R400 - 57', 'R400 - 58', 'R400 - 59', 'R400 - 60', 'R400 - 61', 'R400 - 62', 'R400 - 63', 'R400 - 64'
]

# Economic number mappings
TRAN_ECONOMICO = {
    'R400 - 28 - 1': "28",
    'R400 - 13A ': "13A",
    'R400 - 20A ': "20",
    'R400 - 24A': "24A",
    'R400 - 24B': "24B",
    'R400 - 25A': "25A",
    'R400 - 25B': "25B",
    'R400 - 25C': "25C",
    'R400 - 26A': "26A",
    'R400 - 26B': "26B",
    'R400 - 27A': "27A",
    'R400 - 28A': "28A",
    'R400 - 29A': "29A",
    'R400 - 29B': "29B",
    'R400 - 30A': "30A",
    'R400 - 30B': "30B",
    'R400 - 50 S4': "50",
    'R400 - 101 ': '1A',
    'R400 - 107': '07',
    'R400 - 201': '01',
    'R400 - 202': '02',
    'R400 - 203': '03',
    'R400 - 204': '04',
    'R400 - 205': '05',
    'R400 - 206': '06',
    'R400 - 208': '08',
    'R400 - 209': '09',
    'R400 - 210': '10',
    'R400 - 211': '11',
    'R400 - 230': '23',
    'R400 - 145': '45',
    'R400 - 134': '34',
    'R155 - 28 - 2': '28',
    'R202 - 140': '14A',
    'R202 - 150': '15A',
    'R202 - 170': '17A',
}

# Route configurations
RUTAS = [
    {
        "empresa": "TRANSPORTES PROGRESO, S.A.",
        "ruta": "A6",
        "ramal": "LÓPEZ MATEOS - ESTACIÓN COYOACÁN - ESTACIÓN CHURUBUSCO - CASA BLANCA.",
        "bandera": "A6",
        "unidades": RUTA_A6,
        "km": 84.7,
        "frecuencia": 18,
        "tiempo": 220,
        "file": "A6"
    },
    {
        "empresa": "RUTA 202, S.A. DE C.V.",
        "ruta": "155",
        "ramal": "",
        "bandera": "Apoyo Ruta 155",
        "unidades": RUTA_155,
        "km": 49.0,
        "frecuencia": 10,
        "tiempo": 160,
        "file": "R155"
    },
    {
        "empresa": "RUTA 202, S.A. DE C.V.",
        "ruta": "A202",
        "ramal": "SANTA CATARINA - ESTACIÓN COYOACÁN",
        "bandera": "A202",
        "unidades": RUTA_202,
        "tiempo": 150,
        "km": 62.7,
        "frecuencia": 10,
        "tiempo": 150,
        "file": "R202"
    },
    {
        "empresa": "TRANSPORTACIÓN INTERMUNICIPAL, S.A. DE C.V.",
        "ruta": "31",
        "ramal": "SAN BERNABÉ",
        "bandera": "Ruta 31",
        "unidades": RUTA_31,
        "frecuencia": 15,
        "tiempo": 180,
        "km": 57.2,
        "file": "R31"
    },
    {
        "empresa": "TRANSPORTES PROGRESO, S.A.",
        "ruta": "RUTA SANTA CATARINA 400 ANIVERSARIO",
        "ramal": "PINO SUÁREZ",
        "bandera": "RUTA 400 SECTOR 1",
        "unidades": RUTA_400_S1,
        "frecuencia": 10,
        "tiempo": 130,
        "km": 58.6,
        "file": "R400S1"
    },
    {
        "empresa": "TRANSPORTES PROGRESO, S.A.",
        "ruta": "RUTA SANTA CATARINA 400 ANIVERSARIO",
        "ramal": "FLETEROS",
        "bandera": "RUTA 400 SECTOR 2",
        "unidades": RUTA_400_S2,
        "frecuencia": 10,
        "tiempo": 130,
        "km": 58.6,
        "file": "R400S2"
    }
]

WEEKDAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
MONTHS = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


class SubsidyService:
    """Main service for subsidy operations."""
    
    def __init__(self, user):
        """Initialize the subsidy service."""
        self.user = user
    
    def get_available_routes(self) -> List[Dict[str, Any]]:
        """Get available subsidy routes."""
        return RUTAS
    
    def get_route_units(self, route_code: str) -> List[str]:
        """Get units for a specific route."""
        for ruta in RUTAS:
            if ruta['file'] == route_code:
                return ruta['unidades']
        return []
    
    def get_economic_mapping(self, unit_name: str) -> Optional[str]:
        """Get economic number for a unit."""
        return TRAN_ECONOMICO.get(unit_name)
    
    def create_timesheet_capture(self, date: date, unit_name: str, times: List[List[str]], 
                               driver: str = "", route: str = "") -> TimeSheetCapture:
        """Create a time sheet capture."""
        timesheet, created = TimeSheetCapture.objects.get_or_create(
            date=date,
            name=unit_name,
            defaults={
                'times': times,
                'driver': driver,
                'route': route
            }
        )
        
        if not created:
            timesheet.times = times
            timesheet.driver = driver
            timesheet.route = route
            timesheet.save()
        
        return timesheet
    
    def get_timesheet_data(self, date: date, unit_name: str) -> Optional[TimeSheetCapture]:
        """Get time sheet data for a specific date and unit."""
        try:
            return TimeSheetCapture.objects.get(date=date, name=unit_name)
        except TimeSheetCapture.DoesNotExist:
            return None
    
    def get_ticket_data(self, date: date) -> List[Dict[str, Any]]:
        """Get ticket data for a specific date."""
        # This would integrate with the GPS system to get ticket data
        # For now, return empty list
        return []
    
    def get_available_units(self, date: date) -> List[str]:
        """Get available units for a specific date."""
        # Get units from time sheet captures
        timesheets = TimeSheetCapture.objects.filter(date=date)
        units = list(timesheets.values_list('name', flat=True))
        
        # Add all route units
        for ruta in RUTAS:
            units.extend(ruta['unidades'])
        
        return sorted(list(set(units)))


class TimeSheetGenerator:
    """Generate time sheet reports."""
    
    def __init__(self, user):
        """Initialize the time sheet generator."""
        self.user = user
        self.service = SubsidyService(user)
    
    def generate_timesheet_report(self, route_code: str, start_date: date, 
                                end_date: date, format: str = 'xlsx') -> HttpResponse:
        """Generate time sheet report for a route."""
        # Get route configuration
        route_config = None
        for ruta in RUTAS:
            if ruta['file'] == route_code:
                route_config = ruta
                break
        
        if not route_config:
            raise ValueError(f"Route {route_code} not found")
        
        # Get time sheet data for the date range
        timesheets = TimeSheetCapture.objects.filter(
            date__range=(start_date, end_date),
            name__in=route_config['unidades']
        ).order_by('date', 'name')
        
        if format == 'xlsx':
            return self._generate_excel_report(route_config, timesheets, start_date, end_date)
        elif format == 'csv':
            return self._generate_csv_report(route_config, timesheets, start_date, end_date)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_excel_report(self, route_config: Dict, timesheets, 
                             start_date: date, end_date: date) -> HttpResponse:
        """Generate Excel time sheet report."""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"TimeSheet_{route_config['file']}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Sheet"
        
        # Add headers
        headers = ['Fecha', 'Unidad', 'Conductor', 'Ruta', 'Vueltas', 'Tiempo Total (min)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add data
        row = 2
        for timesheet in timesheets:
            ws.cell(row=row, column=1, value=timesheet.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=timesheet.name)
            ws.cell(row=row, column=3, value=timesheet.driver)
            ws.cell(row=row, column=4, value=timesheet.route)
            ws.cell(row=row, column=5, value=timesheet.rounds_count)
            ws.cell(row=row, column=6, value=timesheet.total_duration)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    def _generate_csv_report(self, route_config: Dict, timesheets, 
                           start_date: date, end_date: date) -> HttpResponse:
        """Generate CSV time sheet report."""
        response = HttpResponse(content_type='text/csv')
        filename = f"TimeSheet_{route_config['file']}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Fecha', 'Unidad', 'Conductor', 'Ruta', 'Vueltas', 'Tiempo Total (min)'])
        
        for timesheet in timesheets:
            writer.writerow([
                timesheet.date.strftime('%Y-%m-%d'),
                timesheet.name,
                timesheet.driver,
                timesheet.route,
                timesheet.rounds_count,
                timesheet.total_duration
            ])
        
        return response


class SubsidyReportGenerator:
    """Generate subsidy reports."""
    
    def __init__(self, user):
        """Initialize the subsidy report generator."""
        self.user = user
        self.service = SubsidyService(user)
    
    def generate_daily_report(self, route_code: str, report_date: date, 
                            format: str = 'xlsx') -> HttpResponse:
        """Generate daily subsidy report."""
        # Get route configuration
        route_config = None
        for ruta in RUTAS:
            if ruta['file'] == route_code:
                route_config = ruta
                break
        
        if not route_config:
            raise ValueError(f"Route {route_code} not found")
        
        # Get daily logs for the route
        daily_logs = DailyLog.objects.filter(
            start__date=report_date,
            route__in=[int(route_code.replace('R', '').replace('A', '')) if route_code.startswith('R') else 0]
        ).order_by('driver__last', 'driver__middle', 'driver__name')
        
        if format == 'xlsx':
            return self._generate_daily_excel_report(route_config, daily_logs, report_date)
        elif format == 'csv':
            return self._generate_daily_csv_report(route_config, daily_logs, report_date)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_daily_excel_report(self, route_config: Dict, daily_logs, 
                                   report_date: date) -> HttpResponse:
        """Generate Excel daily report."""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"DailyReport_{route_config['file']}_{report_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"
        
        # Add title
        title = f"Reporte Diario - {route_config['ruta']} - {report_date.strftime('%d/%m/%Y')}"
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Add headers
        headers = ['Conductor', 'Inicio', 'Fin', 'Ordinarias', 'Preferentes', 'Total', 'A Pagar', 'Pagado', 'Diferencia']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add data
        row = 4
        total_regular = 0
        total_preferent = 0
        total_passengers = 0
        total_due = Decimal('0.00')
        total_payed = Decimal('0.00')
        
        for log in daily_logs:
            ws.cell(row=row, column=1, value=log.driver.full_name)
            ws.cell(row=row, column=2, value=log.start.strftime('%H:%M'))
            ws.cell(row=row, column=3, value=log.stop.strftime('%H:%M'))
            ws.cell(row=row, column=4, value=log.regular)
            ws.cell(row=row, column=5, value=log.preferent)
            ws.cell(row=row, column=6, value=log.total)
            ws.cell(row=row, column=7, value=float(log.due))
            ws.cell(row=row, column=8, value=float(log.payed))
            ws.cell(row=row, column=9, value=float(log.difference))
            
            total_regular += log.regular
            total_preferent += log.preferent
            total_passengers += log.total
            total_due += log.due
            total_payed += log.payed
            row += 1
        
        # Add totals
        ws.cell(row=row, column=1, value="TOTALES")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_regular)
        ws.cell(row=row, column=5, value=total_preferent)
        ws.cell(row=row, column=6, value=total_passengers)
        ws.cell(row=row, column=7, value=float(total_due))
        ws.cell(row=row, column=8, value=float(total_payed))
        ws.cell(row=row, column=9, value=float(total_due - total_payed))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    def _generate_daily_csv_report(self, route_config: Dict, daily_logs, 
                                 report_date: date) -> HttpResponse:
        """Generate CSV daily report."""
        response = HttpResponse(content_type='text/csv')
        filename = f"DailyReport_{route_config['file']}_{report_date.strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(['Conductor', 'Inicio', 'Fin', 'Ordinarias', 'Preferentes', 'Total', 'A Pagar', 'Pagado', 'Diferencia'])
        
        for log in daily_logs:
            writer.writerow([
                log.driver.full_name,
                log.start.strftime('%H:%M'),
                log.stop.strftime('%H:%M'),
                log.regular,
                log.preferent,
                log.total,
                float(log.due),
                float(log.payed),
                float(log.difference)
            ])
        
        return response


class DriverService:
    """Service for driver operations."""
    
    def __init__(self, user):
        """Initialize the driver service."""
        self.user = user
    
    def get_active_drivers(self) -> List[Driver]:
        """Get all active drivers."""
        return Driver.objects.filter(active=True).order_by('last', 'middle', 'name')
    
    def get_driver_by_name(self, full_name: str) -> Optional[Driver]:
        """Get driver by full name."""
        try:
            # Parse full name (format: "Last, First")
            if ',' in full_name:
                last, first = full_name.split(',', 1)
                return Driver.objects.get(last=last.strip(), name=first.strip())
            else:
                return Driver.objects.get(name=full_name)
        except Driver.DoesNotExist:
            return None
    
    def create_driver(self, data: Dict[str, Any]) -> Driver:
        """Create a new driver."""
        return Driver.objects.create(**data)
    
    def update_driver(self, driver_id: int, data: Dict[str, Any]) -> Driver:
        """Update an existing driver."""
        driver = get_object_or_404(Driver, id=driver_id)
        for key, value in data.items():
            setattr(driver, key, value)
        driver.save()
        return driver
    
    def deactivate_driver(self, driver_id: int) -> Driver:
        """Deactivate a driver."""
        driver = get_object_or_404(Driver, id=driver_id)
        driver.active = False
        driver.save()
        return driver 