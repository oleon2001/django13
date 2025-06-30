# coding=utf-8
# Create your views here.

from django.template import RequestContext
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import render_to_response,get_object_or_404, get_list_or_404
from django.db.models import Avg, Max, Min, Q
from django.contrib.auth.decorators import login_required
from django.utils import simplejson
from django.core.urlresolvers import reverse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, DetailView, TemplateView
from django.core.exceptions import ImproperlyConfigured, ObjectDoesNotExist, SuspiciousOperation
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import redirect
import django.core.exceptions
from datetime import timedelta,time,datetime,date

from decimal import Decimal
from django.forms.util import flatatt, ErrorDict, ErrorList

from reportlab.platypus import BaseDocTemplate, PageTemplate, Table, TableStyle, Paragraph, Spacer, Frame, PageBreak
from reportlab.lib.pagesizes import letter,landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle

from django import forms
from django.forms import widgets
from models import *
from views import getPeopleCount,TicketView, dayRangeX

import datetime
from django.views.generic.edit import FormView
from dateutil.parser import parse
from pytz import utc, timezone
import csv
import copy

sdate = datetime.date(2022,7,1)

#RutaA6 = [u'A6 - 02', u'A6 - 03', u'A6 - 04', u'A6 - 05', u'A6 - 06', u'A6 - 09', u'A6 - 11', u'A6 - 14', u'A6 - 18', u'A6 - 21', u'A6 - 22']
RutaA6 = [u'A6 - 01', u'A6 - 02', u'A6 - 03', u'A6 - 04', u'A6 - 05', u'A6 - 06', u'A6 - 07', u'A6 - 08', u'A6 - 09', u'A6 - 10', 
    u'A6 - 11', u'A6 - 12', u'A6 - 13', u'A6 - 14', u'A6 - 15', u'A6 - 16', u'A6 - 17', u'A6 - 18', u'A6 - 19', u'A6 - 20', u'A6 - 21', u'A6 - 22', u'A6 - 26']
Ruta155 = [u'R155 - 01', u'R155 - 02', u'R155 - 03', u'R155 - 04', u'R155 - 05', u'R155 - 06', u'R155 - 07', u'R155 - 11', u'R155 - 12', u'R155 - 13', u'R155 - 14',
  u'R155 - 15', u'R155 - 16', u'R155 - 17', u'R155 - 18', u'R155 - 19', u'R155 - 20', u'R155 - 20A', u'R155 - 21', 
  u'R155 - 22', u'R155 - 23', u'R155 - 24', u'R155 - 25', u'R155 - 26', u'R155 - 27', u'R155 - 28 - 2', u'R155 - 29',
  u'R155 - 30', u'R155 - 31', u'R155 - 70A']
Ruta202 = [u'R202 - 01', u'R202 - 02', u'R202 - 03', u'R202 - 04', u'R202 - 05', u'R202 - 06', u'R202 - 07', u'R202 - 08', u'R202 - 09',
  u'R202 - 10', u'R202 - 11', u'R202 - 12', u'R202 - 13', u'R202 - 14', u'R202 - 140', u'R202 - 15', 
  u'R202 - 16', u'R202 - 150', u'R202 - 17', u'R202 - 170', u'R202 - 18', u'R202 - 19',
  u'R202 - 20', u'R202 - 21', u'R202 - 22', u'R202 - 23', u'R202 - 39']
Ruta207E = [u'R207 - 02', u'R207 - 18', u'R207 - 20', u'R207 - 22', u'R207 - 24', u'R207 - 26', u'R207 - 28', u'R207 - 64', u'R207 - 66']
Ruta207P = [u'R207 - 03', u'R207 - 17', u'R207 - 19', u'R207 - 21', u'R207 - 23', u'R207 - 27', u'R207 - 67', u'R207 - 69']
Ruta31 = [u'R31 - 01', u'R31 - 02', u'R31 - 03', u'R31 - 04', u'R31 - 05', u'R31 - 06', u'R31 - 07', u'R31 - 08', u'R31 - 09', u'R31 - 10',
  u'R31 - 11', u'R31 - 12', u'R31 - 13', u'R31 - 14', u'R31 - 16', u'R31 - 17']
Ruta400S1 = [u'R400 - 01', u'R400 - 02', u'R400 - 03', u'R400 - 04 ', u'R400 - 05', u'R400 - 06', u'R400 - 07', u'R400 - 08',
 u'R400 - 09', u'R400 - 10', u'R400 - 11', u'R400 - 12', u'R400 - 13', u'R400 - 14', u'R400 - 15', u'R400 - 16',
 u'R400 - 17', u'R400 - 18', u'R400 - 19', u'R400 - 20', u'R400 - 21', u'R400 - 22', u'R400 - 23', u'R400 - 24', 
 u'R400 - 25', u'R400 - 26', u'R400 - 27', u'R400 - 28 - 1']
Ruta400S2 = [u'R400 - 32', u'R400 - 33', u'R400 - 34', u'R400 - 35', u'R400 - 36', u'R400 - 37', u'R400 - 38', u'R400 - 39',
 u'R400 - 40', u'R400 - 41', u'R400 - 42', u'R400 - 43', u'R400 - 44', u'R400 - 45', u'R400 - 46', u'R400 - 47',
 u'R400 - 48', u'R400 - 49', u'R400 - 50', u'R400 - 51', u'R400 - 52', u'R400 - 53', u'R400 - 54', u'R400 - 55', u'R400 - 56', u'R400 - 57']
Ruta400S4H = [  u'R400 - 20A ', u'R400 - 62', u'R400 - 63', u'R400 - 70', u'R400 - 71', u'R400 - 76', u'R400 - 107',
 u'R400 - 134', u'R400 - 145', u'R400 - 201', u'R400 - 202', u'R400 - 203', u'R400 - 204', u'R400 - 205',
 u'R400 - 206', u'R400 - 208', u'R400 - 209', u'R400 - 210', u'R400 - 211', u'R400 - 230', u'R400 - 13A',
 u'R400 S4 - 20A ', u'R400 S4 - 62', u'R400 S4 - 63', u'R400 S4 - 70', u'R400 S4 - 71', u'R400 S4 - 76', u'R400 S4 - 107',
 u'R400 S4 - 134', u'R400 S4 - 145', u'R400 S4 - 201', u'R400 S4 - 202', u'R400 S4 - 203', u'R400 S4 - 204', u'R400 S4 - 205',
 u'R400 S4 - 206', u'R400 S4 - 208', u'R400 S4 - 209', u'R400 S4 - 210', u'R400 S4 - 211', u'R400 S4 - 230', u'R400 S4 - 13A',
 ]
Ruta400S4J = [  u'R400 - 24A', u'R400 - 24B', u'R400 - 25A', u'R400 - 25B', u'R400 - 25C', u'R400 - 26A', u'R400 - 26B',
 u'R400 - 27A', u'R400 - 28A', u'R400 - 29A', u'R400 - 29B', u'R400 - 30A', u'R400 - 30B', u'R400 - 50 S4', u'R400 - 101 ',
 u'R400 - 1035', u'R400 - 58', u'R400 - 59', u'R400 - 60', u'R400 - 61', u'R400 - 75', u'R400 - 77', u'R400 - 79',
 u'R400 S4 - 24A', u'R400 S4 - 24B', u'R400 S4 - 25A', u'R400 S4 - 25B', u'R400 S4 - 25C', u'R400 S4 - 26A', u'R400 S4 - 26B',
 u'R400 S4 - 27A', u'R400 S4 - 28A', u'R400 S4 - 29A', u'R400 S4 - 29B', u'R400 S4 - 30A', u'R400 S4 - 30B', u'R400 S4 - 50 S4', u'R400 S4 - 101 ',
 u'R400 S4 - 1035', u'R400 S4 - 58', u'R400 S4 - 59', u'R400 S4 - 60', u'R400 S4 - 61', u'R400 S4 - 75', u'R400 S4 - 77', u'R400 S4 - 79' 
 ]
tran_economico = {
    u'R400 - 28 - 1': u"28",
    u'R400 - 13A ': u"13A",
    u'R400 - 20A ': u"20",
    
    u'R400 - 24A': u"24A",
    u'R400 - 24B': u"24B",
    u'R400 - 25A': u"25A",
    u'R400 - 25B': u"25B",
    u'R400 - 25C': u"25C",
    u'R400 - 26A': u"26A",
    u'R400 - 26B': u"26B",
    u'R400 - 27A': u"27A",
    u'R400 - 28A': u"28A",
    u'R400 - 29A': u"29A",
    u'R400 - 29B': u"29B",
    u'R400 - 30A': u"30A",
    u'R400 - 30B': u"30B",
    u'R400 - 50 S4': u"50",
    
    u'R400 - 101 ': u'1A',
    u'R400 - 107': u'07',
    u'R400 - 201': u'01',
    u'R400 - 202': u'02',
    u'R400 - 203': u'03',
    u'R400 - 204': u'04',
    u'R400 - 205': u'05',
    u'R400 - 206': u'06',
    u'R400 - 208': u'08',
    u'R400 - 209': u'09',
    u'R400 - 210': u'10',
    u'R400 - 211': u'11',
    u'R400 - 230': u'23',
    u'R400 - 145': u'45',
    u'R400 - 134': u'34',
    
    u'R155 - 28 - 2': u'28',
    
    u'R202 - 140': u'14A',
    u'R202 - 150': u'15A',
    u'R202 - 170': u'17A',
    
    }

class DateRangeForm(forms.Form):
	dateStart = forms.DateField(label = "Fecha", widget = widgets.DateInput, initial = datetime.date.today()-timedelta(days=1))
    #dateEnd = forms.DateField(label = "Fecha", widget = widgets.DateInput, initial = datetime.date.today())
	submit_text = "Generar"

	def drange(self):
		d1 = self.cleaned_data['dateStart']
		d1 = datetime.datetime(d1.year,d1.month,d1.day,tzinfo = utc)
		#d2 = self.cleaned_data['dateEnd']
		#d2 = datetime.datetime(d2.year,d2.month,d2.day,tzinfo = utc)
		return (d1,d1+timedelta(hours=1))

class TSheetForm(forms.Form):
    date = forms.DateField(widget = widgets.DateInput, initial = sdate)
    unidad = forms.ChoiceField(choices = (),required = False)
    v1Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v1End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v2Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v2End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v3Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v3End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v4Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v4End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v5Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v5End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v6Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v6End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v7Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v7End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v8Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v8End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v9Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v9End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v10Ini =forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v10End =forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    fromTicket =forms.BooleanField(initial = True, required = False)
    submit_text = "Consultar Unidades"
    
class CsvTSDReportView(FormView):
    template_name = "rutaForm.html"
    form_class = DateRangeForm
    success_url = ''
        
    def form_valid(self,form):
        dt = form.drange()[0]
        rows = TimeSheetCapture.objects.filter(date = dt).order_by("name")

        response = HttpResponse(content_type='text/csv')
        filename = 'Vueltas.%04d.%02d.%02d'%(dt.year,dt.month,dt.day)
        response['Content-Disposition'] = 'attachment; filename="%s.csv"'%filename
        data = []

        for i in rows:
            times = simplejson.loads(i.times)
            r = [i.name]
            tot =0
            for j in times:
                td = parse(j[1]) - parse(j[0])
                mins = (td.seconds %3600) /60
                hours = td.seconds / 3600
                tot +=td.seconds
                r += [j[0],j[1],"{0}:{1:02d}".format(hours,mins)]
            if i.vueltas <6:
                for j in range(i.vueltas,6):
                    r +=["","",""]
            avg = tot / i.vueltas
            mins = (avg %3600) /60
            hours = avg / 3600
            r += [i.vueltas,"{0}:{1:02d}".format(hours,mins) ]
            data.append(r)
        csvwriter = csv.writer(response,dialect="excel")
        csvwriter.writerows(data)
        return response
                
class TsheetView(FormView):
    template_name = "tsheetForm.html"
    form_class = TSheetForm
    success_url = ''

    def form_valid(self,form):
        #ruta = form.cleaned_data['ruta']
        raise ValueError

    def get_context_data(self, **kwargs):
        #raise ValueError
        return super(TsheetView,self).get_context_data(**kwargs)

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        self.request = request
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        return self.render_to_response(self.get_context_data(form=form, stage=0))

    def set_formValues(self,form,times):
        rounds = len(times)
        form.data = form.data.copy()
        #clear all times
        form.data['v1Ini'] = ''
        form.data['v1End'] = ''
        form.data['v2Ini'] = ''
        form.data['v2End'] = ''
        form.data['v3Ini'] = ''
        form.data['v3End'] = ''
        form.data['v4Ini'] = ''
        form.data['v4End'] = ''
        form.data['v5Ini'] = ''
        form.data['v5End'] = ''
        form.data['v6Ini'] = ''
        form.data['v6End'] = ''
        form.data['v7Ini'] = ''
        form.data['v7End'] = ''
        form.data['v8Ini'] = ''
        form.data['v8End'] = ''
        form.data['v9Ini'] = ''
        form.data['v9End'] = ''
        form.data['v10Ini'] = ''
        form.data['v10End'] = ''
        #fill initial values
        if rounds>=1:
            form.data['v1Ini'] = times[0][0]
            form.data['v1End'] = times[0][1]
        if rounds>=2:
            form.data['v2Ini'] = times[1][0]
            form.data['v2End'] = times[1][1]
        if rounds>=3:
            form.data['v3Ini'] = times[2][0]
            form.data['v3End'] = times[2][1]
        if rounds>=4:
            form.data['v4Ini'] = times[3][0]
            form.data['v4End'] = times[3][1]
        if rounds>=5:
            form.data['v5Ini'] = times[4][0]
            form.data['v5End'] = times[4][1]
        if rounds>=6:
            form.data['v6Ini'] = times[5][0]
            form.data['v6End'] = times[5][1]
        if rounds>=7:
            form.data['v7Ini'] = times[6][0]
            form.data['v7End'] = times[6][1]
        if rounds>=8:
            form.data['v8Ini'] = times[7][0]
            form.data['v8End'] = times[7][1]
        if rounds>=9:
            form.data['v9Ini'] = times[8][0]
            form.data['v9End'] = times[8][1]
        if rounds>=10:
            form.data['v10Ini'] = times[9][0]
            form.data['v10End'] = times[9][1]
    
    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        self.request = request
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        # set choices before validation
        try:
            d = form.fields['date'].clean(self.request.POST.get('date',''))
        except:
            return self.render_to_response(self.get_context_data(form = form, stage=0))
        form.fields['date'].widget._is_hidden = True
        #Get tickets from date
        d = datetime.datetime(d.year,d.month,d.day,3)
        tickets = TicketDetails.objects.filter (date__range = (d,d+timedelta(days=1)))
        if not tickets: #invalid date
            return self.render_to_response(self.get_context_data(form = form, stage=0))
        u = 'unidad' in form.data and form.data['unidad']

        # Validate Unidad and get Ticket or ticket choices
        #ctx_units = []
        units = []
        t_contexts = []
        chofer = ""
        for i in tickets:
            context = simplejson.loads(i.ticket)
            #ctx_units.append(context)
            unit = context["Name"]
            if not [unit,unit] in units:
                units.append([unit,unit])
            if unit == u:
                t_contexts.append(context['rounds'])
                chofer = i.chofer
        xtra = RutaA6 + Ruta400S1 + Ruta400S2 + Ruta155 + Ruta202
        for i in xtra:
            if not [i,i] in units:
                units.append([i,i])
            if (u == i) and not t_contexts:
                t_contexts = True
        if not t_contexts:
            units.sort()
            form.fields['unidad'].choices = units
            form.submit_text = "Importar Ticket"
            return self.render_to_response(self.get_context_data(form = form, stage=1))

        if not isinstance(t_contexts,list): #t_contexts == True:
            t_contexts = []
        form.fields['unidad'].choices = [[u,u]]
        form.fields['unidad'].widget._is_hidden = True

        #Check if we have it in the DB
        date = datetime.datetime(d.year,d.month,d.day,tzinfo = utc)
        unidad = u
        if 'fromTicket' in form.data:
            recs = None
        else:
            recs = TimeSheetCapture.objects.filter(date = date, name = unidad)
        stage = 2
        if recs:
            times = simplejson.loads(recs[0].times)
            times.sort()
            rounds = len(times)
            form.submit_text = "Modificar"
            stage = 3
        else:
            if 'fromTicket' in form.data:
                form.submit_text = "Modificar"
            else:
                form.submit_text = "Agregar"
            times = []
            rounds = 0 #len(ticket)
            for t in t_contexts:
                for i in t:
                    r = [i['start'], i['stop']]
                    if not r in times:
                        times.append(r)
                        rounds +=1
            times = times[0:10]
            times.sort()
        if not 'v1Ini' in form.data:
            saveFlag = False
            self.set_formValues(form,times)
        else:
            saveFlag = True
        if form.is_valid():
            if saveFlag:
                form.submit_text = "Modificar"
                stage = 3
                #get times array
                times = []
                if form.cleaned_data['v1Ini'] and form.cleaned_data['v1End']:
                    times.append([form.cleaned_data['v1Ini'].strftime("%H:%M"),form.cleaned_data['v1End'].strftime("%H:%M")])
                if form.cleaned_data['v2Ini'] and form.cleaned_data['v2End']:
                    times.append([form.cleaned_data['v2Ini'].strftime("%H:%M"),form.cleaned_data['v2End'].strftime("%H:%M")])
                if form.cleaned_data['v3Ini'] and form.cleaned_data['v3End']:
                    times.append([form.cleaned_data['v3Ini'].strftime("%H:%M"),form.cleaned_data['v3End'].strftime("%H:%M")])
                if form.cleaned_data['v4Ini'] and form.cleaned_data['v4End']:
                    times.append([form.cleaned_data['v4Ini'].strftime("%H:%M"),form.cleaned_data['v4End'].strftime("%H:%M")])
                if form.cleaned_data['v5Ini'] and form.cleaned_data['v5End']:
                    times.append([form.cleaned_data['v5Ini'].strftime("%H:%M"),form.cleaned_data['v5End'].strftime("%H:%M")])
                if form.cleaned_data['v6Ini'] and form.cleaned_data['v6End']:
                    times.append([form.cleaned_data['v6Ini'].strftime("%H:%M"),form.cleaned_data['v6End'].strftime("%H:%M")])
                if form.cleaned_data['v7Ini'] and form.cleaned_data['v7End']:
                    times.append([form.cleaned_data['v7Ini'].strftime("%H:%M"),form.cleaned_data['v7End'].strftime("%H:%M")])
                if form.cleaned_data['v8Ini'] and form.cleaned_data['v8End']:
                    times.append([form.cleaned_data['v8Ini'].strftime("%H:%M"),form.cleaned_data['v8End'].strftime("%H:%M")])
                if form.cleaned_data['v9Ini'] and form.cleaned_data['v9End']:
                    times.append([form.cleaned_data['v9Ini'].strftime("%H:%M"),form.cleaned_data['v9End'].strftime("%H:%M")])
                if form.cleaned_data['v10Ini'] and form.cleaned_data['v10End']:
                    times.append([form.cleaned_data['v10Ini'].strftime("%H:%M"),form.cleaned_data['v10End'].strftime("%H:%M")])
                times.sort()
                #refill values
                self.set_formValues(form,times)
                timejson = simplejson.dumps(times)
                if recs:
                    rec = recs[0]
                    rec.vueltas = len(times)
                    rec.times = timejson
                    if rec.vueltas:
                        rec.save()
                    else:
                        rec.delete()
                else:
                    rec = TimeSheetCapture(name = unidad, date = date, chofer = chofer, vueltas = len(times), times = timejson)
                    rec.save()
            return self.render_to_response(self.get_context_data(form = form, stage=stage))
        else:
            return self.render_to_response(self.get_context_data(form = form, stage=stage))
            

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

base_dir = '/home/django13/skyguard/gps/tracker/'
templates = base_dir + 'templates/'
output = base_dir + 'reportes/'
weekdays = ["Lunes","Martes","Miércoles", "Jueves", "Viernes","Sábado","Domingo"]
months = ["","Enero","Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
Rutas = [
    { "empresa":    u"TRANSPORTES PROGRESO, S.A.",
      "ruta":       u"A6",
      "ramal":      u"LÓPEZ MATEOS - ESTACIÓN COYOACÁN - ESTACIÓN CHURUBUSCO - CASA BLANCA.",
      "bandera":    u"A6",
      "unidades":   RutaA6,
      "km":         84.7, #84.720,
      "frecuencia": 18,
      "tiempo":     220,
      "workbook":   None,
      "diario":     [],
      "file":       "A6"},
    { "empresa":    u"RUTA 202, S.A. DE C.V.",
      "ruta":       u"155",
      "ramal":      u"",
      "bandera":    u"Apoyo Ruta 155",
      "unidades":   Ruta155,
      "km":         49.0, #49.017,
      "frecuencia": 10,
      "tiempo":     160,
      "workbook":   None,
      "diario":     [],
      "file":       "R155"},
    { "empresa":    u"RUTA 202, S.A. DE C.V.",
      "ruta":       u"A202",
      "ramal":      u"SANTA CATARINA - ESTACIÓN COYOACÁN",
      "bandera":    u"A202",
      "unidades":   Ruta202,
      "tiempo":     150,
      "km":         62.7, #62.734,
      "frecuencia": 10,
      "tiempo":     150,
      "workbook":   None,
      "diario":     [],
      "file":       "R202"},
    { "empresa":    u"TRANSPORTACIÓN INTERMUNICIPAL, S.A. DE C.V.",
      "ruta":       u"31",
      "ramal":      u"SAN BERNABÉ",
      "bandera":    u"Ruta 31",
      "unidades":   Ruta31,
      "frecuencia": 15,
      "tiempo":     180,
      "km":         57.2, #57.204,
      "workbook":   None,
      "diario":     [],
      "file":       "R31"},
    { "empresa":    u"TRANSPORTACIÓN INTERMUNICIPAL, S.A. DE C.V.",
      "ruta":       u"207",
      "ramal":      u"ESCOBEDO",
      "bandera":    u"Ruta 207",
      "unidades":   Ruta207E,
      "frecuencia": 22,
      "tiempo":     135,
      "km":         41.5, #41.485,
      "workbook":   None,
      "diario":     [],
      "file":       "R207E"},
    { "empresa":    u"TRANSPORTACIÓN INTERMUNICIPAL, S.A. DE C.V.",
      "ruta":       u"207",
      "ramal":      u"PENITENCIARÍA",
      "bandera":    u"Ruta 207",
      "unidades":   Ruta207P,
      "frecuencia": 22,
      "tiempo":     135,
      "km":         39.7, #39.693,
      "workbook":   None,
      "diario":     [],
      "file":       "R207P"},
    { "empresa":    u"TRANSPORTES PROGRESO, S.A.",
      "ruta":       u"RUTA SANTA CATARINA 400 ANIVERSARIO",
      "ramal":      u"PINO SUÁREZ",
      "bandera":    u"RUTA 400 SECTOR 1",
      "unidades":   Ruta400S1,
      "frecuencia": 8,
      "tiempo":     125,
      "km":         41.4, #41.359,
      "workbook":   None,
      "diario":     [],
      "file":       "R400S1"},
    { "empresa":    u"TRANSPORTES PROGRESO, S.A.",
      "ruta":       u"RUTA SANTA CATARINA 400 ANIVERSARIO",
      "ramal":      u"FLETEROS",
      "bandera":    u"RUTA 400 SECTOR 2",
      "unidades":   Ruta400S2,
      "frecuencia": 7,
      "tiempo":     130,
      "km":         40.7, #40.712,
      "workbook":   None,
      "diario":     [],
      "file":       "R400S2"},
    { "empresa":    u"TRANSPORTES PROGRESO, S.A.",
      "ruta":       u"RUTA SANTA CATARINA 400 ANIVERSARIO",
      "ramal":      u"SECTOR 4 HÉROES - CERÁMICA",
      "bandera":    u"RUTA 400 SECTOR 4 HÉROES",
      "unidades":   Ruta400S4H,
      "km":         50.8, #50.843,
      "frecuencia": 10,
      "tiempo":     110,
      "workbook":   None,
      "diario":     [],
      "file":       "R400S4H"},
    { "empresa":    u"TRANSPORTES PROGRESO, S.A.",
      "ruta":       u"RUTA SANTA CATARINA 400 ANIVERSARIO",
      "ramal":      u"SECTOR 4 SAN JOSÉ - CERÁMICA",
      "bandera":    u"RUTA 400 SECTOR 4 SAN JOSÉ",
      "unidades":   Ruta400S4J,
      "frecuencia": 10,
      "tiempo":     130,
      "km":         58.6, #58.589,
      "workbook":   None,
      "diario":     [],
      "file":       "R400S4J"},      
    ]

def loadWorkbooks(start,smonth,stop,endmonth,year):
    for i in Rutas:
        i["workbook"] = openpyxl.load_workbook(templates+"Template.xlsx")
        i["diario"] = []
        wb = i["workbook"]
        wb[u"Día"]["C7"].value = i["ruta"]
        wb[u"Día"]["C8"].value = i["ramal"]
        wb[u"Día"]["D9"].value = i["bandera"]
        wb[u"Resumen"]["A1"].value = i["empresa"]
        wb[u"Resumen"]["B4"].value = i["ruta"]
        wb[u"Resumen"]["B5"].value = i["ramal"]
        wb[u"Resumen"]["B6"].value = i["bandera"]
        wb[u"Resumen"]["B7"].value = i["km"]
        wb[u"Resumen"]["C8"].value = u"del día {0} de {1} al día {2} de {3} de {4}".format(start,months[smonth],stop,months[endmonth],year).upper()
        for j in [u'4T', u'5T', u'6T', u'7T', u'8T', u'9T']:
            wb[j]['B5'] = i["empresa"]
            wb[j]['B6'] = i["ruta"]
            wb[j]['B7'] = i["ramal"]
        wb[u'4T']['M7'] = i["bandera"]
        wb[u'5T']['P7'] = i["bandera"]
        wb[u'6T']['S7'] = i["bandera"]
        wb[u'7T']['U7'] = i["bandera"]
        wb[u'8T']['Y7'] = i["bandera"]
        wb[u'9T']['AB7'] = i["bandera"]
        
def addDetail(ruta,day,regs):
    gold = PatternFill(patternType = 'solid', fgColor = openpyxl.styles.colors.Color(rgb = "FFFF00"))
    
    r = []
    vmax = 4
    #fill economico
    regs = list(regs)
    for i in regs:
        if i.name in tran_economico:
            i.econ = tran_economico[i.name]
        else: 
            i.econ = i.name.split('-')[-1].strip()
    # sort by economico
    regs.sort(key = lambda a: a.econ)
    for i in regs:
        if i.name in ruta["unidades"]:
            i.itimes = []
            vtas = simplejson.loads(i.times)
            if i.vueltas > vmax:
                vmax = i.vueltas
            for j in vtas:
                t0 = parse(j[0])
                t1 = parse(j[1])
                #dt = t1-t0
                it0 = t0.hour/24.0 +t0.minute/1440.0
                it1 = t1.hour/24.0 +t1.minute/1440.0
                if t1.hour<4:
                    idt = 1+it1-it0
                else:
                    idt = it1-it0 #dt.seconds / 86400.0
                i.itimes.append([it0,it1,idt])
            r.append(i)
    if not r:
        print "No data fount for {0} @ {1}.".format(ruta["file"],day.strftime("%d-%m-%Y"))
        #return #####-- Add sheet in any case--#####
    if vmax > 9: 
        raise ValueError("Too many rounds for {0} @ {1}".format(ruta["file"],day.strftime("%d-%m-%Y")))
    # ADD Detail Worksheet
    ws = ruta["workbook"].copy_worksheet(ruta["workbook"][u"{0}T".format(vmax)])
    ws.title = u"Detalle {0} {1}".format(months[day.month],day.day)
    halfs = []
    
    row = 12
    col = 1
    vtot = 0
    ttot = 0
    ht1 = (ruta["tiempo"]/2-27)/1440.0
    ht2 = (ruta["tiempo"]/2+27)/1440.0
    if not r:
        for i in range(2,16):
            ws.cell(row=12,column = i).value = 0
    for i in r:
        ws.cell(row = row, column =1).value = i.econ
        tot = 0.0
        v =0
        for j in i.itimes:
            ws.cell(row = row, column = 2+v*3).value = j[0]
            ws.cell(row = row, column = 3+v*3).value = j[1]
            ws.cell(row = row, column = 4+v*3).value = j[2]
            tot += j[2]
            v +=1
            
        if (ruta == Rutas[-1]) or (ruta == Rutas[-2]): #Sector 4
            # if (i.vueltas >1) and (i.itimes[0][2] <= ht2) and (i.itimes[-1][2] <= ht2):
                # ws.cell(row = row, column =4).fill = gold
                # ws.cell(row = row, column =1+len(i.itimes)*3).fill = gold
                # ws['A40'].value = "Aclaración: El sector 4 normalmente inicia y termina con medias vueltas. Se ajustó el total de vueltas de las unidades marcadas en este color."
                # for j in range(1,11):
                    # ws.cell(row = 40, column =j).fill = gold
                # i.vueltas -=1
            if (i.vueltas >=1) and ((i.itimes[0][2] <= ht2) or (i.itimes[-1][2] <= ht2)):
                if i.itimes[0][2] <= ht2:
                    ws.cell(row = row, column =4).fill = gold
                    i.vueltas -= 0.5
                if i.itimes[-1][2] <= ht2:
                    ws.cell(row = row, column =1+len(i.itimes)*3).fill = gold
                    i.vueltas -= 0.5
                ws['A40'].value = u"Aclaración: El sector 4 normalmente inicia y termina con medias vueltas. Se ajustó el total de vueltas de las unidades marcadas en este color."
                for j in range(1,11):
                    ws.cell(row = 40, column =j).fill = gold
                if i.vueltas == int(i.vueltas):
                    i.vueltas = int(i.vueltas)
                else:
                    halfs.append(i.econ)
        ws.cell(row = row, column =vmax*3+2).value = i.vueltas
        ws.cell(row = row, column =vmax*3+3).value = i.vueltas and tot/i.vueltas
        vtot += i.vueltas
        ttot +=tot
        row +=1
    if halfs:
        disc = None
        if len(halfs)%2 == 1:
            disc = halfs.pop()
        if halfs:
            n = len(halfs)/2
            if n>1:
                v = str(n)+' vueltas.'
            else:
                v = str(n)+' vuelta.'            
            ws['A41'] = u"Aclaración: Las medias vueltas de los camiones "+','.join(halfs)+u' se tomaron en cuenta en pares para sumar '+v
            for j in range(1,11):
                ws.cell(row = 41, column =j).fill = gold
            if disc:
                ws['A42'] = u"Aclaración: La media vuelta del camión "+disc+u" se descartó." 
                for j in range(1,11):
                    ws.cell(row = 42, column =j).fill = gold
        else:
            ws['A41'] = u"Aclaración: La media vuelta del camión "+disc+u" se descartó." 
            for j in range(1,11):
                ws.cell(row = 41, column =j).fill = gold
    vtot = int(vtot)
    ws.cell(row = 39, column = vmax*3+2).value = vtot
    ws.cell(row = 39, column = vmax*3+3).value = float(vtot and ttot/vtot)
    ws["B8"].value = day.strftime("%d/%m/%Y")
    ws["E8"].value = weekdays[day.weekday()]
    ruta["diario"].append({"day":day, "vtot": vtot, "unidades": len(r), "tavg": float(vtot and ttot/vtot), "vmax": vmax, "data": r})
    
    # ADD Summary Worksheet
    ws = ruta["workbook"].copy_worksheet(ruta["workbook"][u"Día"])
    ws.title = u"Resumen {0} {1}".format(months[day.month],day.day)
    ws["E15"].value = len(r) #i["unidades"]
    ws["E16"].value = int((vtot and ttot/vtot)*1440) #int(i["tavg"]*1440)
    ws["E17"].value = len(r) and ruta["frecuencia"]
    ws["E18"].value = len(r) and ruta["km"]
    ws["E19"].value = vtot #i["vtot"]
    ws["E20"].value = float(len(r) and vtot/len(r)) #float(i["vtot"])/i["unidades"]
    ws["G7"].value = day.strftime("%d/%m/%Y") #i["day"].strftime("%d/%m/%Y")
    #

def HiLiteSheet(ruta):
    colors = [
            PatternFill(patternType = 'solid', fgColor = openpyxl.styles.colors.Color(rgb = "DDEBF7")),
            PatternFill(patternType = 'solid', fgColor = openpyxl.styles.colors.Color(rgb = "BDD7EE")),
            PatternFill(patternType = 'solid', fgColor = openpyxl.styles.colors.Color(rgb = "9BC2E6")),
            PatternFill(patternType = 'solid', fgColor = openpyxl.styles.colors.Color(rgb = "2F75B5"))
        ]
    gold = PatternFill(patternType = 'solid', fgColor = openpyxl.styles.colors.Color(rgb = "FFFF00"))
    orange = PatternFill(patternType = 'solid', fgColor = openpyxl.styles.colors.Color(rgb = "F4B084"))
    for data in ruta["diario"]:
        day = data["day"]
        r = data["data"]
        vmax = data['vmax']
        ws = ruta['workbook'][u"Detalle {0} {1}".format(months[day.month],day.day)]
        # get repeat round times
        repeats = {}
        for i in r: 
         for j in i.itimes:
            if j[2] in repeats: 
                if repeats[j[2]] <4: repeats[j[2]] +=1
            else: repeats[j[2]] = 0
        row = 12
        starts = {}
        for i in r:
            for j in range(len(i.itimes)):
                # same times in day
                if i.itimes[j][0] in starts:
                    ws.cell(row = row, column = 2+j*3).fill = gold
                    ws.cell(row = starts[i.itimes[j][0]][0], column = 2+3*starts[i.itimes[j][0]][1]).fill = gold
                else:
                    starts[i.itimes[j][0]]=(row,j)
                ###
                #get deviation from avgf
                dev = abs(i.itimes[j][2] - ruta['tiempo']/1440.0)  #abs(i.itimes[j][2] - data['tavg'])  #/data['tavg']
                # round times
                if dev > 45.0/1440.0: #0.40:
                    ws.cell(row = row, column = 4+j*3).fill = orange
                elif repeats[i.itimes[j][2]]:
                    ws.cell(row = row, column = 4+j*3).fill = colors[repeats[i.itimes[j][2]]-1]
                # overlaps
                if j and ((i.itimes[j-1][1]+(0.5/1440.0)) >= i.itimes[j][0]):
                    ws.cell(row = row, column = j*3).fill   = gold
                    ws.cell(row = row, column = 2+j*3).fill = gold
            row+=1
   
def addSummary(ruta):
    SumRow = 11
    mxnVueltas = 0
    thin = Side(border_style="thin", color="000000")
    border = Border(top= thin, left = thin, bottom = thin, right = thin)
    ali = Alignment(horizontal="center", vertical="bottom")
    nf0 = 'General'
    nf1 = '#,##0' # or '#,##0.0'
    nf2 = '_-$* #,##0.00_-;-$* #,##0.00_-' #Formato anterior '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
    print "Summary for ", ruta["ramal"]," - ",len(ruta["diario"])
    for i in ruta["diario"]:
        ws = ruta["workbook"]["Resumen"]
        ws.insert_rows(SumRow)
        ws.cell(row = SumRow, column =1).value = "{0} {1}".format(months[i["day"].month],i["day"].day)
        ws.cell(row = SumRow, column =2).value = i["vtot"]
        ws.cell(row = SumRow, column =3).value = int(i["vtot"]*ruta["km"])
        ws.cell(row = SumRow, column =4).value = int(i["vtot"]*ruta["km"])*2
        mxnVueltas += int(i["vtot"]*ruta["km"])*2
        ws.cell(row = SumRow, column =1).border = border
        ws.cell(row = SumRow, column =2).border = border
        ws.cell(row = SumRow, column =3).border = border
        ws.cell(row = SumRow, column =4).border = border
        ws.cell(row = SumRow, column =1).alignment = ali
        ws.cell(row = SumRow, column =2).alignment = ali
        ws.cell(row = SumRow, column =3).alignment = ali
        ws.cell(row = SumRow, column =4).alignment = ali
        ws.cell(row = SumRow, column =1).number_format = nf0
        ws.cell(row = SumRow, column =2).number_format = nf0
        ws.cell(row = SumRow, column =3).number_format = nf1
        ws.cell(row = SumRow, column =4).number_format = nf2
        SumRow +=1
    ws.cell(row = SumRow, column =4).value = mxnVueltas
   
def generate_xls(start,end,month,year):
    stop = datetime.datetime(year,month,start,tzinfo = utc)+datetime.timedelta(days=end+1-start)
    loadWorkbooks(start,month,stop.day-1,stop.month,year) #Corrected display of the last day
    for d in range (0,end+1-start):
        day = datetime.datetime(year,month,start,tzinfo = utc)+datetime.timedelta(days=d)
        regs = TimeSheetCapture.objects.filter(date = day).order_by("name")
        for r in Rutas:
            addDetail(r,day,regs)
        print "Day {0} done.".format(d)
    for r in Rutas:
        addSummary(r)
    print "Saving files."
    for r in Rutas:
        sheetnames = [u'D\xeda', u'4T', u'5T', u'6T', u'7T', u'8T', u'9T']
        for i in sheetnames:
            r["workbook"].remove(r["workbook"][i])
        fname ="{0} {1} {2}-{3}.xlsx".format(r["file"],months[month],start,end)
        print "Saving ",fname
        r["workbook"].save(output+fname)
    print "Saving hints."
    for r in Rutas:
        HiLiteSheet(r)
        fname ="Hint {0} {1} {2}-{3}.xlsx".format(r["file"],months[month],start,end)
        print "Saving ",fname
        r["workbook"].save(output+fname)
    