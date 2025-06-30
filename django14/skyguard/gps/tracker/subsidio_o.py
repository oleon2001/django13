# coding=utf-8
# Create your views here.

from django.template import RequestContext
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import render_to_response,get_object_or_404, get_list_or_404
from django.db.models import Avg, Max, Min, Q
from django.contrib.auth.decorators import login_required
from django.utils import simplejson
from django.core.urlresolvers import reverse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, DetailView, TemplateView
from django.core.exceptions import ImproperlyConfigured, ObjectDoesNotExist, SuspiciousOperation
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import redirect
import django.core.exceptions
from datetime import timedelta,time,datetime,date

from decimal import Decimal
from django.forms.util import flatatt, ErrorDict, ErrorList

from reportlab.platypus import BaseDocTemplate, PageTemplate, Table, TableStyle, Paragraph, Spacer, Frame, PageBreak
from reportlab.lib.pagesizes import letter,landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle

from django import forms
from django.forms import widgets
from models import *
from views import getPeopleCount,TicketView, dayRangeX

import datetime
from django.views.generic.edit import FormView
from dateutil.parser import parse
from pytz import utc, timezone
import csv

sdate = datetime.date(2022,6,1)

class DateRangeForm(forms.Form):
	dateStart = forms.DateField(label = "Fecha", widget = widgets.DateInput, initial = datetime.date.today()-timedelta(days=1))
    #dateEnd = forms.DateField(label = "Fecha", widget = widgets.DateInput, initial = datetime.date.today())
	submit_text = "Generar"

	def drange(self):
		d1 = self.cleaned_data['dateStart']
		d1 = datetime.datetime(d1.year,d1.month,d1.day,tzinfo = utc)
		#d2 = self.cleaned_data['dateEnd']
		#d2 = datetime.datetime(d2.year,d2.month,d2.day,tzinfo = utc)
		return (d1,d1+timedelta(hours=1))

class TSheetForm(forms.Form):
    date = forms.DateField(widget = widgets.DateInput, initial = sdate)
    unidad = forms.ChoiceField(choices = (),required = False)
    v1Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v1End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v2Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v2End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v3Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v3End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v4Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v4End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v5Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v5End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v6Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v6End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v7Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v7End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v8Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v8End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v9Ini = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v9End = forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v10Ini =forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    v10End =forms.TimeField(input_formats = ["%H:%M"], initial ="" ,required = False)
    submit_text = "Consultar Unidades"
    
class CsvTSDReportView(FormView):
    template_name = "rutaForm.html"
    form_class = DateRangeForm
    success_url = ''
        
    def form_valid(self,form):
        dt = form.drange()[0]
        rows = TimeSheetCapture.objects.filter(date = dt).order_by("name")

        response = HttpResponse(content_type='text/csv')
        filename = 'Vueltas.%04d.%02d.%02d'%(dt.year,dt.month,dt.day)
        response['Content-Disposition'] = 'attachment; filename="%s.csv"'%filename
        data = []

        for i in rows:
            times = simplejson.loads(i.times)
            r = [i.name]
            tot =0
            for j in times:
                td = parse(j[1]) - parse(j[0])
                mins = (td.seconds %3600) /60
                hours = td.seconds / 3600
                tot +=td.seconds
                r += [j[0],j[1],"{0}:{1:02d}".format(hours,mins)]
            if i.vueltas <6:
                for j in range(i.vueltas,6):
                    r +=["","",""]
            avg = tot / i.vueltas
            mins = (avg %3600) /60
            hours = avg / 3600
            r += [i.vueltas,"{0}:{1:02d}".format(hours,mins) ]
            data.append(r)
        csvwriter = csv.writer(response,dialect="excel")
        csvwriter.writerows(data)
        return response
                
class TsheetView(FormView):
    template_name = "tsheetForm.html"
    form_class = TSheetForm
    success_url = ''

    def form_valid(self,form):
        #ruta = form.cleaned_data['ruta']
        raise ValueError

    def get_context_data(self, **kwargs):
        #raise ValueError
        return super(TsheetView,self).get_context_data(**kwargs)

    @method_decorator(login_required)
    def get(self, request, *args, **kwargs):
        self.request = request
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        return self.render_to_response(self.get_context_data(form=form, stage=0))

    def set_formValues(self,form,times):
        rounds = len(times)
        form.data = form.data.copy()
        #clear all times
        form.data['v1Ini'] = ''
        form.data['v1End'] = ''
        form.data['v2Ini'] = ''
        form.data['v2End'] = ''
        form.data['v3Ini'] = ''
        form.data['v3End'] = ''
        form.data['v4Ini'] = ''
        form.data['v4End'] = ''
        form.data['v5Ini'] = ''
        form.data['v5End'] = ''
        form.data['v6Ini'] = ''
        form.data['v6End'] = ''
        form.data['v7Ini'] = ''
        form.data['v7End'] = ''
        form.data['v8Ini'] = ''
        form.data['v8End'] = ''
        form.data['v9Ini'] = ''
        form.data['v9End'] = ''
        form.data['v10Ini'] = ''
        form.data['v10End'] = ''
        #fill initial values
        if rounds>=1:
            form.data['v1Ini'] = times[0][0]
            form.data['v1End'] = times[0][1]
        if rounds>=2:
            form.data['v2Ini'] = times[1][0]
            form.data['v2End'] = times[1][1]
        if rounds>=3:
            form.data['v3Ini'] = times[2][0]
            form.data['v3End'] = times[2][1]
        if rounds>=4:
            form.data['v4Ini'] = times[3][0]
            form.data['v4End'] = times[3][1]
        if rounds>=5:
            form.data['v5Ini'] = times[4][0]
            form.data['v5End'] = times[4][1]
        if rounds>=6:
            form.data['v6Ini'] = times[5][0]
            form.data['v6End'] = times[5][1]
        if rounds>=7:
            form.data['v7Ini'] = times[6][0]
            form.data['v7End'] = times[6][1]
        if rounds>=8:
            form.data['v8Ini'] = times[7][0]
            form.data['v8End'] = times[7][1]
        if rounds>=9:
            form.data['v9Ini'] = times[8][0]
            form.data['v9End'] = times[8][1]
        if rounds>=10:
            form.data['v10Ini'] = times[9][0]
            form.data['v10End'] = times[9][1]
    
    @method_decorator(login_required)
    def post(self, request, *args, **kwargs):
        self.request = request
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        # set choices before validation
        try:
            d = form.fields['date'].clean(self.request.POST.get('date',''))
        except:
            return self.render_to_response(self.get_context_data(form = form, stage=0))
        form.fields['date'].widget._is_hidden = True
        #Get tickets from date
        d = datetime.datetime(d.year,d.month,d.day,3)
        tickets = TicketDetails.objects.filter (date__range = (d,d+timedelta(days=1)))
        if not tickets: #invalid date
            return self.render_to_response(self.get_context_data(form = form, stage=0))
        u = 'unidad' in form.data and form.data['unidad']

        # Validate Unidad and get Ticket or ticket choices
        #ctx_units = []
        units = []
        t_contexts = []
        for i in tickets:
            context = simplejson.loads(i.ticket)
            #ctx_units.append(context)
            unit = context["Name"]
            if not [unit,unit] in units:
                units.append([unit,unit])
            if unit == u:
                t_contexts.append(context['rounds'])
                chofer = i.chofer
        if not t_contexts:
            units.sort()
            form.fields['unidad'].choices = units
            form.submit_text = "Importar Ticket"
            return self.render_to_response(self.get_context_data(form = form, stage=1))

        form.fields['unidad'].choices = [[u,u]]
        form.fields['unidad'].widget._is_hidden = True

        #Check if we have it in the DB
        date = datetime.datetime(d.year,d.month,d.day,tzinfo = utc)
        unidad = u
        recs = TimeSheetCapture.objects.filter(date = date, name = unidad)
        stage = 2
        if recs:
            times = simplejson.loads(recs[0].times)
            times.sort()
            rounds = len(times)
            form.submit_text = "Modificar"
            stage = 3
        else:
            form.submit_text = "Agregar"
            times = []
            rounds = 0 #len(ticket)
            for t in t_contexts:
                for i in t:
                    r = [i['start'], i['stop']]
                    if not r in times:
                        times.append(r)
                        rounds +=1
            times = times[0:10]
            times.sort()
        if not 'v1Ini' in form.data:
            saveFlag = False
            self.set_formValues(form,times)
        else:
            saveFlag = True
        if form.is_valid():
            if saveFlag:
                form.submit_text = "Modificar"
                stage = 3
                #get times array
                times = []
                if form.cleaned_data['v1Ini'] and form.cleaned_data['v1End']:
                    times.append([form.cleaned_data['v1Ini'].strftime("%H:%M"),form.cleaned_data['v1End'].strftime("%H:%M")])
                if form.cleaned_data['v2Ini'] and form.cleaned_data['v2End']:
                    times.append([form.cleaned_data['v2Ini'].strftime("%H:%M"),form.cleaned_data['v2End'].strftime("%H:%M")])
                if form.cleaned_data['v3Ini'] and form.cleaned_data['v3End']:
                    times.append([form.cleaned_data['v3Ini'].strftime("%H:%M"),form.cleaned_data['v3End'].strftime("%H:%M")])
                if form.cleaned_data['v4Ini'] and form.cleaned_data['v4End']:
                    times.append([form.cleaned_data['v4Ini'].strftime("%H:%M"),form.cleaned_data['v4End'].strftime("%H:%M")])
                if form.cleaned_data['v5Ini'] and form.cleaned_data['v5End']:
                    times.append([form.cleaned_data['v5Ini'].strftime("%H:%M"),form.cleaned_data['v5End'].strftime("%H:%M")])
                if form.cleaned_data['v6Ini'] and form.cleaned_data['v6End']:
                    times.append([form.cleaned_data['v6Ini'].strftime("%H:%M"),form.cleaned_data['v6End'].strftime("%H:%M")])
                if form.cleaned_data['v7Ini'] and form.cleaned_data['v7End']:
                    times.append([form.cleaned_data['v7Ini'].strftime("%H:%M"),form.cleaned_data['v7End'].strftime("%H:%M")])
                if form.cleaned_data['v8Ini'] and form.cleaned_data['v8End']:
                    times.append([form.cleaned_data['v8Ini'].strftime("%H:%M"),form.cleaned_data['v8End'].strftime("%H:%M")])
                if form.cleaned_data['v9Ini'] and form.cleaned_data['v9End']:
                    times.append([form.cleaned_data['v9Ini'].strftime("%H:%M"),form.cleaned_data['v9End'].strftime("%H:%M")])
                if form.cleaned_data['v10Ini'] and form.cleaned_data['v10End']:
                    times.append([form.cleaned_data['v10Ini'].strftime("%H:%M"),form.cleaned_data['v10End'].strftime("%H:%M")])
                times.sort()
                #refill values
                self.set_formValues(form,times)
                timejson = simplejson.dumps(times)
                if recs:
                    rec = recs[0]
                    rec.vueltas = len(times)
                    rec.times = timejson
                    rec.save()
                else:
                    rec = TimeSheetCapture(name = unidad, date = date, chofer = chofer, vueltas = len(times), times = timejson)
                    rec.save()
            return self.render_to_response(self.get_context_data(form = form, stage=stage))
        else:
            return self.render_to_response(self.get_context_data(form = form, stage=stage))
            

import openpyxl
base_dir = '/home/django13/skyguard/gps/tracker/'
templates = base_dir + 'templates/'
output = base_dir + 'reportes/'
RutaA6 = [u'A6 - 02', u'A6 - 03', u'A6 - 04', u'A6 - 05', u'A6 - 06', u'A6 - 09', u'A6 - 11', u'A6 - 14', u'A6 - 18', u'A6 - 21', u'A6 - 22']
Ruta155 = [u'R155 - 02', u'R155 - 06', u'R155 - 07', u'R155 - 12', u'R155 - 13', u'R155 - 14', u'R155 - 15', u'R155 - 16', u'R155 - 17',
  u'R155 - 18', u'R155 - 19', u'R155 - 20', u'R155 - 25', u'R155 - 28 - 2', u'R155 - 31']
Ruta202 = [u'R202 - 01', u'R202 - 02', u'R202 - 03', u'R202 - 04', u'R202 - 05', u'R202 - 06', u'R202 - 07', u'R202 - 08', u'R202 - 09',
  u'R202 - 10', u'R202 - 11', u'R202 - 12', u'R202 - 13', u'R202 - 14', u'R202 - 15', u'R202 - 150', u'R202 - 17', u'R202 - 18', u'R202 - 19',
  u'R202 - 20', u'R202 - 21', u'R202 - 22', u'R202 - 23', u'R202 - 39']
Ruta207E = [u'R207 - 02', u'R207 - 18', u'R207 - 20', u'R207 - 22', u'R207 - 24', u'R207 - 26', u'R207 - 28', u'R207 - 64', u'R207 - 66']
Ruta207P = [u'R207 - 03', u'R207 - 17', u'R207 - 19', u'R207 - 21', u'R207 - 23', u'R207 - 27', u'R207 - 67', u'R207 - 69']
Ruta31 = [u'R31 - 01', u'R31 - 02', u'R31 - 03', u'R31 - 04', u'R31 - 05', u'R31 - 06', u'R31 - 07', u'R31 - 08', u'R31 - 09', u'R31 - 10',
  u'R31 - 11', u'R31 - 12', u'R31 - 13', u'R31 - 14', u'R31 - 16', u'R31 - 17']
Ruta400S1 = [u'R400 - 01', u'R400 - 02', u'R400 - 04 ', u'R400 - 05', u'R400 - 07', u'R400 - 08', u'R400 - 13', u'R400 - 15', u'R400 - 16',
 u'R400 - 17', u'R400 - 18', u'R400 - 19', u'R400 - 28 - 1']
Ruta400S2 = [u'R400 - 27', u'R400 - 34', u'R400 - 35', u'R400 - 37', u'R400 - 39', u'R400 - 41', u'R400 - 44', u'R400 - 45', u'R400 - 46',
 u'R400 - 49', u'R400 - 54', u'R400 - 56']
Ruta400S4H = [  u'R400 - 20A ', u'R400 - 62', u'R400 - 63', u'R400 - 70', u'R400 - 71', u'R400 - 76', u'R400 - 107',
 u'R400 - 134', u'R400 - 145', u'R400 - 201', u'R400 - 202', u'R400 - 203', u'R400 - 204', u'R400 - 205',
 u'R400 - 206', u'R400 - 208', u'R400 - 209', u'R400 - 210', u'R400 - 211', u'R400 - 230', u'R400 - 13A']

Ruta400S4J = [  u'R400 - 24A', u'R400 - 24B', u'R400 - 25A', u'R400 - 25B', u'R400 - 25C', u'R400 - 26A', u'R400 - 26B',
 u'R400 - 27A', u'R400 - 28A', u'R400 - 29A', u'R400 - 29B', u'R400 - 30A', u'R400 - 30B', u'R400 - 50 S4', u'R400 - 101 ',
 u'R400 - 1035', u'R400 - 58', u'R400 - 59', u'R400 - 60', u'R400 - 61', u'R400 - 75', u'R400 - 77', u'R400 - 79']


tran_economico = {
    u'R400 - 28 - 1': u"28",
    u'R400 - 13A ': u"130",
    u'R400 - 20A ': u"120",
    
    u'R400 - 24A': u"2400",
    u'R400 - 24B': u"124",
    u'R400 - 25A': u"2500",
    u'R400 - 25B': u"25",
    u'R400 - 25C': u"125",
    u'R400 - 26A': u"260",
    u'R400 - 26B': u"126",
    u'R400 - 27A': u"270",
    u'R400 - 28A': u"280",
    u'R400 - 29A': u"290",
    u'R400 - 29B': u"400",
    u'R400 - 30A': u"300",
    u'R400 - 30B': u"2300",
    u'R400 - 50 S4': u"150",
    }
    
def getRegs(regs,ruta):
    r = []
    vmax = 0
    for i in regs:
        if i.name in ruta: 
            i.itimes = []
            if i.name in tran_economico:
                i.econ = tran_economico[i.name]
            else: 
                i.econ = i.name.split('-')[-1].strip()
            vtas = simplejson.loads(i.times)
            r.append(i)
            if i.vueltas > vmax:
                vmax = i.vueltas
            for j in vtas:
                t0 = parse(j[0])
                t1 = parse(j[1])
                dt = t1-t0
                it0 = t0.hour/24.0 +t0.minute/1440.0
                it1 = t1.hour/24.0 +t1.minute/1440.0
                idt = dt.seconds / 86400.0
                i.itimes.append([it0,it1,idt])
    if   vmax <= 4: 
        tname = "4T.xlsx"
        vmax = 4
    elif vmax == 5: tname = "5T.xlsx"
    elif vmax == 6: tname = "6T.xlsx"
    elif vmax == 7: tname = "7T.xlsx"
    elif vmax == 8: tname = "8T.xlsx"
    elif vmax == 9: tname = "9T.xlsx"
    else: raise ValueError("Demasiadas vueltas")
    tot_col = vmax*3+2
    return r,tname,tot_col
            
def fillXls(ws, ruta, last_col,meters):
    row = 12
    col = 1
    vtot = 0
    ttot = 0
    for i in ruta:
        ws.cell(row = row, column =1).value = i.econ
        tot = 0.0
        v =0
        for j in i.itimes:
            ws.cell(row = row, column = 2+v*3).value = j[0]
            ws.cell(row = row, column = 3+v*3).value = j[1]
            ws.cell(row = row, column = 4+v*3).value = j[2]
            tot += j[2]
            v +=1
        ws.cell(row = row, column =last_col).value = i.vueltas
        ws.cell(row = row, column =last_col+1).value = tot/i.vueltas
        vtot += i.vueltas
        ttot +=tot
        row +=1
    ws.cell(row = 39, column = last_col).value = vtot
    ws.cell(row = 39, column = last_col+1).value = ttot/vtot
    return len(ruta),vtot
    

weekdays = ["Lunes","Martes","Miércoles", "Jueves", "Viernes","Sábado","Domingo"]
months = ["","Enero","Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

Headers = [
    [u"Ruta A6.%Y-%m-%d.xlsx",u"TRANSPORTE PROGRESO, S.A.","A6",u"LÓPEZ MATEOS - ESTACIÓN COYOACÁN - ESTACIÓN CHURUBUSCO - CASA BLANCA.","A6",RutaA6,84720,u"Diario A6.%Y-%m-%d.xlsx",[],u"Mensual A6.%Y-%m-%d.xlsx"],
    [u"Ruta 155.%Y-%m-%d.xlsx",u"RUTA 202, S.A. DE C.V.","155",u"",u"Apoyo Ruta 155",Ruta155,49017,u"Diario 155.%Y-%m-%d.xlsx",[],"Mensual 155.%Y-%m-%d.xlsx"],
    [u"Ruta 202.%Y-%m-%d.xlsx",u"RUTA 202, S.A. DE C.V.",u"A202","SANTA CATARINA - ESTACIÓN COYOACÁN",u"A202",Ruta202,62734,"Diario 202.%Y-%m-%d.xlsx",[],u"Mensual 202.%Y-%m-%d.xlsx"],
    [u"Ruta 31.%Y-%m-%d.xlsx",u"TRANSPORTACIÓN INTERMUNICIPAL, S.A. DE C.V.",u"31",u"SAN BERNABÉ",u"RUTA 31",Ruta31,57204,u"Diario 31.%Y-%m-%d.xlsx",[],u"Mensual 31.%Y-%m-%d.xlsx"],
    [u"Ruta 207E.%Y-%m-%d.xlsx",u"TRANSPORTACIÓN INTERMUNICIPAL, S.A. DE C.V.","207",u"ESCOBEDO",u"RUTA 207",Ruta207E,41485,u"Diario 207E.%Y-%m-%d.xlsx",[],u"Mensual 207E.%Y-%m-%d.xlsx"],
    [u"Ruta 207P.%Y-%m-%d.xlsx",u"TRANSPORTACIÓN INTERMUNICIPAL, S.A. DE C.V.","207",u"PENITENCIARÍA",u"RUTA 207",Ruta207P,39693,u"Diario 207P.%Y-%m-%d.xlsx",[],u"Mensual 207P.%Y-%m-%d.xlsx"],
    [u"Ruta 400S1.%Y-%m-%d.xlsx",u"TRANSPORTE PROGRESO, S.A.",u"RUTA STANTA CATARINA 400 ANIVERSARIO",u"PINO SUÁREZ",u"RUTA 400 SECTOR 1",Ruta400S1,41359,"uDiario 400S1.%Y-%m-%d.xlsx",[],u"Mensual 400S1.%Y-%m-%d.xlsx"],
    [u"Ruta 400S2.%Y-%m-%d.xlsx",u"TRANSPORTE PROGRESO, S.A.",u"RUTA STANTA CATARINA 400 ANIVERSARIO",u"FLETEROS",u"RUTA 400 SECTOR 2",Ruta400S2,40712,u"Diario 400S2.%Y-%m-%d.xlsx",[],u"Mensual 400S2.%Y-%m-%d.xlsx"],
    [u"Ruta 400S4H.%Y-%m-%d.xlsx",u"TRANSPORTE PROGRESO, S.A.",u"RUTA STANTA CATARINA 400 ANIVERSARIO",u"HÉROES",u"RUTA 400 SECTOR 4 HÉROES",Ruta400S4H,50843,"uDiario 400S4H.%Y-%m-%d.xlsx",[],u"Mensual 400S4H.%Y-%m-%d.xlsx"],
    [u"Ruta 400S4J.%Y-%m-%d.xlsx",u"TRANSPORTE PROGRESO, S.A.",u"RUTA STANTA CATARINA 400 ANIVERSARIO",u"SAN JOSÉ",u"RUTA 400 SECTOR 4 SAN JOSE",Ruta400S4J,58589,"uDiario 400S4J.%Y-%m-%d.xlsx",[],u"Mensual 400S4J.%Y-%m-%d.xlsx"],
    ]
 
def fillHeaders(wb,hdrs,day,last_col,nunits,vtot):
    fname = day.strftime(hdrs[0])
    wb.active['B5'].value = hdrs[1]
    wb.active['B6'].value = hdrs[2]
    wb.active['B7'].value = hdrs[3]
    wb.active['B8'].value = day.strftime("%d/%m/%Y")
    wb.active['E8'].value = weekdays[day.weekday()]
    wb.active.cell(row = 7, column = last_col-1).value = hdrs[4]
    wb.save(output+fname)

    fname = day.strftime(hdrs[7])
    wb = openpyxl.load_workbook(templates+"Diario.xlsx")
    wb.active['C7'].value = hdrs[2]
    wb.active['C8'].value = hdrs[3]
    wb.active['D9'].value = hdrs[4]
    wb.active['G7'].value = day.strftime("%d/%m/%Y")
    wb.active['E15'].value = nunits
    wb.active['E18'].value = hdrs[6]/1000.0
    wb.active['E19'].value = vtot
    wb.active['E20'].value = vtot*1.0/nunits
    wb.save(output+fname)
    hdrs[8].append([day, vtot, vtot *hdrs[6]/1000.0, vtot * 2 * hdrs[6]/1000])

from copy import copy

def copyCell(shd,row,col,source_cell):
    target_cell = shd.cell(column=col, row=row)
    target_cell._value = source_cell._value
    target_cell.data_type = source_cell.data_type

    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

    #if source_cell.hyperlink:
    #    target_cell._hyperlink = copy(source_cell.hyperlink)

    #if source_cell.comment:
    #    target_cell.comment = copy(source_cell.comment)


def copyDiario(wbo,rows):
    sho = wbo.active
    wb = openpyxl.Workbook()
    shd = wb.active
    shd.title ="RESUMEN DIARIO"
    shd.sheet_format = copy(sho.sheet_format)
    shd.sheet_properties = copy(sho.sheet_properties)
    shd.merged_cells = copy(sho.merged_cells)
    shd.page_margins = copy(sho.page_margins)
    shd.freeze_panes = copy(sho.freeze_panes)

    total = 0;
    for (row, col), source_cell in sho._cells.items():
        if row < 11: copyCell(shd,row,col,source_cell)
        elif row == 11:
            n = 0
            for r in rows[8]:
                if col ==1: source_cell.value = r[0].strftime("%d/%m/%Y")
                if col ==2: source_cell.value = r[1]
                if col ==3: source_cell.value = r[2]
                if col ==4: 
                    source_cell.value = r[3]
                    total = total + r[3]
                copyCell(shd,11+n,col,source_cell)
                n+=1
        elif row ==12:
            if col==4: 
                source_cell.value = total
            copyCell(shd,11+len(rows[8]),col,source_cell)
        else: copyCell(shd, row+len(rows[8])-1, col, source_cell)
    shd['A1'].value = rows[1]
    shd['B4'].value = rows[2]
    shd['B5'].value = rows[3]
    shd['B6'].value = rows[4]
    shd['B7'].value = rows[6]/1000.0
    shd['A8'].value = "PERIODO DE ESTIMACION: DEL DIA {0} DE {1} AL DIA {2} DE {3} DE {4}".format(
        rows[8][0][0].day, months[rows[8][0][0].month],rows[8][-1][0].day,months[rows[8][-1][0].month],rows[8][0][0].year)
        
    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(sho.row_dimensions)):
        shd.row_dimensions[rn] = copy(sho.row_dimensions[rn])
    if sho.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        shd.sheet_format.defaultColWidth = copy(sho.sheet_format.defaultColWidth)
    for key, value in sho.column_dimensions.items():
        shd.column_dimensions[key].min = copy(sho.column_dimensions[key].min)  
        shd.column_dimensions[key].max = copy(sho.column_dimensions[key].max)  
        shd.column_dimensions[key].width = copy(sho.column_dimensions[key].width)
        shd.column_dimensions[key].hidden = copy(sho.column_dimensions[key].hidden)
    wb.save(output+rows[8][0][0].strftime(rows[9]))

    
def generate_xls(day,ndays):
    mes = months[day.month]
    dstart = day
    for d in range(ndays):
        regs = TimeSheetCapture.objects.filter(date = day).order_by("name")
        if not regs: raise ValueError("No data found")
        for i in Headers:
            ruta,tname, last_col = getRegs(regs,i[5])
            if ruta:
                wb = openpyxl.load_workbook(templates+tname)
                nunits,vtot = fillXls(wb.active,ruta,last_col,i[6])
                fillHeaders(wb,i,day,last_col,nunits,vtot)
            else:
                print "No data found for "+i[4]+ day.strftime(" %d-%m-%Y")
        day += timedelta(days=1)
    for h in Headers:
        wb = openpyxl.load_workbook(templates+"Periodo.xlsx")
        copyDiario(wb,h)
    